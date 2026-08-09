"""Canonical reference data.

Loaded once from the authoritative source files (never hand-maintained lists):

* ``HS_with_ALL_descriptions_AND_UNITS.xlsx`` — 6.5k official 11-digit Nepal HS
  codes + tariff UNIT(S).  Authoritative for HS validity and supplementary unit.
* ``countries-code.xlsx``                     — canonical Alpha-2 catalogue.
* ``bank_code_and_bank_names_with_swift_code.csv`` — 279 Nepal banks + SWIFT.
* ``terms_of_payments_and_codes.csv``         — ASYCUDA payment-term codes.
* ``customs_offices.csv``       — 71 NECAS office codes (NNSW codification DB,
  scraped 2026-08-01: nnsw.gov.np/trade/help/codification, reference=office).
* ``declaration_models.csv``    — the 17 Box-1 type+code lines (IM 4, EX 1, …).
* ``extended_procedures.csv``   — NECAS ANNEX 1 (55 rows; 9100 is duplicated in
  the source annex with two different descriptions — both rows kept, pending
  DoC verification; code-validity lookups only need membership).
* ``national_procedures.csv``   — NECAS ANNEX 3 (182 rows; the numbering gaps
  at 330/361/363/374/375 are in the source — validate against the explicit
  set, never against ranges).
* ``transport_modes.csv``       — Box 25/26 modes 01–09.
* ``incoterms.csv``             — Incoterms 2020 codes for the delivery-terms
  dropdown ("C&F" from legacy documents is accepted as an escape value, not
  listed here).

The store exposes deterministic lookups.  It is the *only* HS/COO/bank/terms/
office/procedure authority; the LLM never finalises any of these.
"""
from __future__ import annotations

import csv
import math
import re
import unicodedata
from dataclasses import dataclass, field
from functools import lru_cache

import openpyxl

from ..config import REFERENCE_DIR, get_settings


# --------------------------------------------------------------------------- #
# Normalisation helpers
# --------------------------------------------------------------------------- #
def _norm_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def digits_only(s: str | None) -> str:
    return re.sub(r"\D", "", s or "")


# ---- HS search query handling --------------------------------------------- #
_NUMERICISH = re.compile(r"^[\d\s.,\-/]+$")     # digit query with common separators
_HS_TOKEN_RE = re.compile(r"[a-z0-9]+")


def hs_text_tokens(text: str) -> tuple[str, ...]:
    """Unicode- and case-normalized alphanumeric tokens (search index form)."""
    s = unicodedata.normalize("NFKD", text or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return tuple(_HS_TOKEN_RE.findall(s.lower()))


#

# No goods description a reviewer searches for is longer than this; the cap is
# here so an oversized query cannot be used to scale the full-tariff scan.
_MAX_HS_QUERY_CHARS = 120


def hs_query_error(query: str) -> str | None:
    """Reject queries too short to search usefully (None = valid)."""
    q = (query or "").strip()
    if len(q) < 2:
        return "Search query must be at least 2 characters."
    if len(q) > _MAX_HS_QUERY_CHARS:
        return f"Search query must be {_MAX_HS_QUERY_CHARS} characters or fewer."
    if _NUMERICISH.fullmatch(q) and len(digits_only(q)) < 2:
        return "Numeric HS search needs at least 2 digits."
    return None


# Controlled country aliases (COO_country_of_origin_rules.txt).  NA -> Namibia,
# never a null marker.
COUNTRY_ALIASES = {
    "uk": "GB", "u.k.": "GB", "unitedkingdom": "GB", "britain": "GB", "greatbritain": "GB",
    "usa": "US", "u.s.a.": "US", "unitedstates": "US", "unitedstatesofamerica": "US", "america": "US",
    "uae": "AE", "u.a.e.": "AE", "unitedarabemirates": "AE",
    "turkiye": "TR", "turkey": "TR",
    "cotedivoire": "CI", "ivorycoast": "CI",
    "southkorea": "KR", "korea": "KR", "republicofkorea": "KR", "rok": "KR",
    "northkorea": "KP", "russia": "RU", "russianfederation": "RU",
    "vietnam": "VN", "prc": "CN", "china": "CN", "hongkong": "HK", "taiwan": "TW",
    "na": "NA", "namibia": "NA",
}

# Payment-term aliases / patterns.  Never silently default unknown to LC
# (ADR-006).  Tenor drafts ("75 DAYS FROM AWB DATE") are parsed separately.
# Short aliases (<=3 chars) match only as whole words so "DA" never fires
# inside "DAYS" and "DAP" never inside a longer token.
TERMS_ALIASES = {
    "lc": "200", "l/c": "200", "letterofcredit": "200", "documentarycredit": "200",
    "tt": "400", "t/t": "400", "telegraphictransfer": "400", "banktransfer": "400", "wiretransfer": "400",
    "dap": "600", "documentagainstpayment": "600", "documentsagainstpayment": "600", "dp": "600",
    "cad": "700", "cashagainstdocuments": "700", "advancepayment": "700", "advance": "700",
    "da": "906", "documentsagainstacceptance": "906", "documentagainstacceptance": "906",
    "draftagainstacceptance": "906",
    "foc": "999", "freeofcharge": "999",
    "barter": "300",
}


@dataclass
class HsRecord:
    hs11: str
    description: str
    unit: str
    # Display-only supporting text (workbook column "AI Generated Explanation &
    # Included Products").  NEVER classification authority, NEVER a ranking or
    # search input — the official description is the only searchable text.
    explanation: str = ""


@dataclass
class BankRecord:
    code: str
    name: str
    swift: str


@dataclass
class DeclarationModelRecord:
    """One Box-1 line — the reviewer picks a LINE, never type and code apart
    (IM 4 exists, IM 1 does not)."""
    type: str          # IM / EX / MIS / PEI / PEM / PEX / TIV
    code: str          # single digit, e.g. "4"
    description: str


@dataclass
class ReferenceStore:
    hs_by_11: dict[str, HsRecord] = field(default_factory=dict)
    hs_by_prefix: dict[str, list[str]] = field(default_factory=dict)  # 8/6/4-digit -> [hs11]
    country_by_norm: dict[str, str] = field(default_factory=dict)     # normalised name -> alpha2
    valid_alpha2: set[str] = field(default_factory=set)
    country_name_by_alpha2: dict[str, str] = field(default_factory=dict)
    bank_by_swift: dict[str, BankRecord] = field(default_factory=dict)
    bank_by_norm_name: dict[str, BankRecord] = field(default_factory=dict)
    banks: list[BankRecord] = field(default_factory=list)
    terms_by_code: dict[str, str] = field(default_factory=dict)
    # ---- regime / office / transport references (per-job selection) -------- #
    office_by_code: dict[str, str] = field(default_factory=dict)      # code -> name
    declaration_models: list[DeclarationModelRecord] = field(default_factory=list)
    declaration_model_pairs: set[tuple[str, str]] = field(default_factory=set)
    # ordered annex rows for the dropdowns (9100 appears twice, both kept) …
    extended_procedures: list[tuple[str, str]] = field(default_factory=list)
    national_procedures: list[tuple[str, str]] = field(default_factory=list)
    # … and membership dicts for validation (first description wins on 9100 —
    # a DELIBERATE dedup, membership is all validation needs)
    extended_proc_by_code: dict[str, str] = field(default_factory=dict)
    national_proc_by_code: dict[str, str] = field(default_factory=dict)
    transport_mode_by_code: dict[str, str] = field(default_factory=dict)
    incoterm_by_code: dict[str, str] = field(default_factory=dict)
    # hs11 -> normalized official-description tokens, built once at load time
    # so interactive search never re-reads or re-tokenizes the workbook
    hs_desc_tokens: dict[str, tuple[str, ...]] = field(default_factory=dict)
    # hs11 -> AI-explanation tokens (display-only column, used as a weaker
    # signal for the whole-DB semantic HS match)
    hs_expl_tokens: dict[str, tuple[str, ...]] = field(default_factory=dict)
    # token -> set of hs11 codes carrying it (description OR explanation) — an
    # inverted index so best_hs_by_description scans only relevant candidates
    hs_token_index: dict[str, set[str]] = field(default_factory=dict)

    # ---- HS -------------------------------------------------------------- #
    def hs_exact(self, hs11: str) -> HsRecord | None:
        return self.hs_by_11.get(digits_only(hs11).zfill(11) if digits_only(hs11) else "")

    def hs_candidates_for_prefix(self, prefix: str) -> list[HsRecord]:
        codes = self.hs_by_prefix.get(prefix, [])
        return [self.hs_by_11[c] for c in codes]

    def hs_unit(self, hs11: str) -> str | None:
        rec = self.hs_exact(hs11)
        return rec.unit if rec else None

    def best_hs_by_description(self, description: str) -> tuple[HsRecord, float] | None:
        """Best whole-DB semantic match for a goods description — used to
        finalize an HS when the invoice prints no HS hint (user rule
        2026-07-19).  Deterministic token overlap: a word in the official
        DESCRIPTION counts full, a word only in the AI explanation counts half;
        the inverted index limits the scan to codes that share a token.
        Returns (record, 0-100 score) or None when no word overlaps at all."""
        q = [t for t in hs_text_tokens(description) if len(t) >= 2 and not t.isdigit()]
        if not q:
            return None
        # Rare words are far more discriminative than common ones ("winding"
        # nails the code; "of"/"other" do not).  Weight each match by IDF =
        # log(1 + N/df) so a distinctive word dominates the score — this is what
        # keeps a description on the code that shares its SPECIFIC term.
        n_docs = len(self.hs_by_11) or 1
        idf = {t: math.log(1.0 + n_docs / (1 + len(self.hs_token_index.get(t, ()))))
               for t in set(q)}
        denom = sum(idf[t] for t in q) or 1.0
        cand: set[str] = set()
        for t in q:
            cand |= self.hs_token_index.get(t, set())
        best_code, best_score = None, 0.0
        for code in cand:
            desc = self.hs_desc_tokens.get(code, ())
            expl = self.hs_expl_tokens.get(code, ())
            total = 0.0
            for w in q:
                if w in desc or any(len(w) >= 4 and d.startswith(w) for d in desc):
                    total += idf[w]
                elif w in expl or any(len(w) >= 4 and e.startswith(w) for e in expl):
                    total += 0.5 * idf[w]
            score = total / denom
            if score > best_score or (score == best_score and (best_code is None or code < best_code)):
                best_score, best_code = score, code
        if not best_code or best_score <= 0:
            return None
        return self.hs_by_11[best_code], round(100.0 * best_score, 1)

    def search_hs(self, query: str, limit: int = 30) -> list[tuple[HsRecord, float]]:
        """Deterministic local search over the official HS database.

        Numeric queries match by digit prefix (exact 11-digit first); text
        queries use token similarity over the OFFICIAL DESCRIPTION only — the
        display-only explanation column is never a ranking input.  Every query
        token must match (minimum relevance threshold).  Sort: score
        descending, hs11 ascending.  No I/O, no network.
        """
        q = (query or "").strip()
        limit = max(1, min(50, int(limit)))
        # The text branch below is a query-token x every-HS-description-token
        # cross-product over the whole tariff (~12k codes).  The RESULT was
        # capped but the WORK was not, so a long query multiplied the scan by
        # its token count — one authenticated request able to spend arbitrary
        # CPU, repeatable.  Bound the query itself: 12 tokens is far past any
        # real goods description, and the extra tokens only ever narrowed an
        # already-narrow match.
        qtokens_cap = 12
        scored: list[tuple[float, str]] = []
        if _NUMERICISH.fullmatch(q):
            digits = digits_only(q)
            for code in self.hs_by_11:
                if code == digits:
                    scored.append((100.0, code))          # exact above all prefixes
                elif code.startswith(digits):
                    scored.append((90.0, code))
        else:
            qtokens = hs_text_tokens(q)[:qtokens_cap]
            for code, dtokens in self.hs_desc_tokens.items():
                total = 0.0
                for qt in qtokens:
                    best = 0.0
                    for dt in dtokens:
                        if dt == qt:
                            best = 1.0
                            break
                        if dt.startswith(qt):
                            best = max(best, 0.8)
                        elif len(qt) >= 4 and qt in dt:
                            best = max(best, 0.6)
                    if best == 0.0:                        # every token must match
                        total = -1.0
                        break
                    total += best
                if total > 0 and qtokens:
                    scored.append((round(100.0 * total / len(qtokens), 1), code))
        scored.sort(key=lambda t: (-t[0], t[1]))
        return [(self.hs_by_11[c], s) for s, c in scored[:limit]]

    # ---- Countries ------------------------------------------------------- #
    def normalize_country(self, raw: str | None) -> str | None:
        if not raw:
            return None
        s = raw.strip()
        # already an alpha-2?
        if len(s) == 2 and s.upper() in self.valid_alpha2:
            return s.upper()
        key = _norm_text(s)
        if not key:
            return None
        if key in COUNTRY_ALIASES:
            code = COUNTRY_ALIASES[key]
            return code if code in self.valid_alpha2 else None
        return self.country_by_norm.get(key)

    # ---- Banks ----------------------------------------------------------- #
    def resolve_bank(self, swift: str | None, name: str | None) -> tuple[BankRecord | None, str]:
        if swift:
            base = swift.strip().upper()
            for cand in (base, base[:8], base[:8] + "XXX"):
                if cand in self.bank_by_swift:
                    return self.bank_by_swift[cand], "SWIFT"
            # base-8 match against stored 8/11 char codes
            b8 = base[:8]
            for rec in self.banks:
                if rec.swift[:8] == b8 and b8:
                    return rec, "SWIFT_BASE8"
        if name:
            key = _norm_text(name)
            if key in self.bank_by_norm_name:
                return self.bank_by_norm_name[key], "NAME_EXACT"
            for rec in self.banks:  # conservative containment
                nk = _norm_text(rec.name)
                if nk and (nk in key or key in nk) and len(key) > 5:
                    return rec, "NAME_FUZZY"
        return None, "UNRESOLVED"

    # ---- Payment terms --------------------------------------------------- #
    def resolve_terms_code(self, raw: str | None) -> tuple[str | None, str]:
        if not raw:
            return None, "MISSING"
        key = _norm_text(raw)
        if not key:
            return None, "MISSING"
        tokens: set[str] = set()
        for t in re.split(r"[^A-Za-z0-9/]+", str(raw)):
            if not t:
                continue
            tokens.add(_norm_text(t))                    # "L/C" -> "lc"
            tokens.update(_norm_text(p) for p in t.split("/") if p)  # "LC/TT" -> "lc","tt"
        for alias, code in TERMS_ALIASES.items():
            norm = _norm_text(alias)
            if len(norm) <= 3:
                if norm in tokens:          # whole-word only: "DA" must not match "DAYS"
                    return code, "ALIAS"
            elif norm in key:
                return code, "ALIAS"
        # exact code appears in the text
        for code in self.terms_by_code:
            if code in digits_only(raw):
                return code, "CODE"
        return None, "REVIEW_REQUIRED"

    def terms_description(self, code: str) -> str:
        return self.terms_by_code.get(code, "")


# --------------------------------------------------------------------------- #
# Loading
# --------------------------------------------------------------------------- #
def _load_hs(store: ReferenceStore) -> None:
    path = get_settings().hs_excel_path
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c or "").strip().upper() for c in row]
            continue
        if not row or row[0] in (None, ""):
            continue
        code_raw = str(row[0]).strip()
        code = digits_only(code_raw).zfill(11)
        if len(code) != 11:
            continue
        desc = str(row[1] or "").strip()
        unit = str(row[2] or "").strip().upper() or "UNT"
        explanation = str(row[3] or "").strip() if len(row) > 3 else ""
        store.hs_by_11[code] = HsRecord(hs11=code, description=desc, unit=unit,
                                        explanation=explanation)
        dtok = hs_text_tokens(desc)
        etok = hs_text_tokens(explanation)
        store.hs_desc_tokens[code] = dtok
        store.hs_expl_tokens[code] = etok
        for t in {w for w in (*dtok, *etok) if len(w) >= 2 and not w.isdigit()}:
            store.hs_token_index.setdefault(t, set()).add(code)
        for n in (8, 6, 4):
            store.hs_by_prefix.setdefault(code[:n], []).append(code)
    wb.close()


def _load_countries(store: ReferenceStore) -> None:
    path = REFERENCE_DIR / "countries-code.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    first = True
    for name, alpha2 in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if not alpha2:
            continue
        code = str(alpha2).strip().upper()
        nm = str(name or "").strip()
        store.valid_alpha2.add(code)
        store.country_name_by_alpha2[code] = nm
        store.country_by_norm[_norm_text(nm)] = code
    wb.close()


def _load_banks(store: ReferenceStore) -> None:
    path = get_settings().bank_csv_path
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            code = str(r.get("Bank Code") or "").strip()
            name = str(r.get("Bank Name") or "").strip()
            swift = str(r.get("SWIFT Code") or "").strip().upper()
            if not code:
                continue
            rec = BankRecord(code=code, name=name, swift=swift)
            store.banks.append(rec)
            if swift:
                store.bank_by_swift[swift] = rec
            store.bank_by_norm_name[_norm_text(name)] = rec


def _load_terms(store: ReferenceStore) -> None:
    path = get_settings().terms_csv_path
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            code = str(r.get("Code") or "").strip()
            desc = str(r.get("Description") or "").strip()
            if code:
                store.terms_by_code[code] = desc


def _read_code_rows(filename: str, value_column: str) -> list[tuple[str, str]]:
    """Ordered (code, value) rows of a two-column reference CSV.  Hard
    ``REFERENCE_DIR`` path (the countries pattern): these files are reference
    data, not deployment configuration."""
    path = REFERENCE_DIR / filename
    rows: list[tuple[str, str]] = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            code = str(r.get("Code") or "").strip()
            if code:
                rows.append((code, str(r.get(value_column) or "").strip()))
    return rows


def _load_regime_references(store: ReferenceStore) -> None:
    store.office_by_code = dict(_read_code_rows("customs_offices.csv", "Name"))

    path = REFERENCE_DIR / "declaration_models.csv"
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.DictReader(fh):
            typ = str(r.get("Type") or "").strip().upper()
            code = str(r.get("Code") or "").strip()
            if typ and code:
                store.declaration_models.append(DeclarationModelRecord(
                    type=typ, code=code, description=str(r.get("Description") or "").strip()))
                store.declaration_model_pairs.add((typ, code))

    store.extended_procedures = _read_code_rows("extended_procedures.csv", "Description")
    for code, desc in store.extended_procedures:
        store.extended_proc_by_code.setdefault(code, desc)   # 9100: first row wins
    store.national_procedures = _read_code_rows("national_procedures.csv", "Description")
    store.national_proc_by_code = dict(store.national_procedures)
    store.transport_mode_by_code = dict(_read_code_rows("transport_modes.csv", "Description"))
    store.incoterm_by_code = dict(_read_code_rows("incoterms.csv", "Description"))


@lru_cache
def get_reference() -> ReferenceStore:
    store = ReferenceStore()
    _load_hs(store)
    _load_countries(store)
    _load_banks(store)
    _load_terms(store)
    _load_regime_references(store)
    return store
